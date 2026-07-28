import openpyxl
from core.models import Brand, Product

def import_glossary_from_excel(file_path):
    """
    Lee la pestaña 'Glosario códigos' e importa las Marcas y Productos en la BD.
    """
    print(f"Abriendo archivo Excel: {file_path}")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    sheet_name = 'Glosario códigos'
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"No se encontró la pestaña '{sheet_name}' en el Excel.")
    
    ws = wb[sheet_name]
    
    products_created = 0
    brands_created = 0
    
    # Recorremos las filas omitiendo las filas de encabezado (comenzando en fila 6)
    for r_idx in range(5, ws.max_row + 1):
        item_code = ws.cell(row=r_idx, column=2).value # Col B: Código
        origin = ws.cell(row=r_idx, column=3).value    # Col C: Origen
        family_name = ws.cell(row=r_idx, column=4).value # Col D: Familia (Marca)
        subfamily = ws.cell(row=r_idx, column=5).value  # Col E: Subfamilia
        description = ws.cell(row=r_idx, column=6).value # Col F: Descripción
        truck_mounted_str = ws.cell(row=r_idx, column=7).value # Col G: Sobre camión
        
        # Validamos que al menos tengamos código y descripción para procesar
        if not item_code or not description:
            continue
            
        # Normalizamos la marca (Familia)
        brand_name = str(family_name).strip() if family_name else "GENERIC"
        brand_code = brand_name.upper()
        
        # Obtenemos o creamos la Marca
        brand, created = Brand.objects.get_or_create(
            code=brand_code,
            defaults={'name': brand_name, 'is_active': True}
        )
        if created:
            brands_created += 1
            print(f"Nueva marca registrada: {brand_name}")
            
        # Normalizamos el origen (Venta / Compra)
        origin_val = 'VENTA'
        if origin and 'compra' in str(origin).lower():
            origin_val = 'COMPRA'
            
        # Normalizamos sobre camión (Sí/No)
        is_truck_mounted = False
        if truck_mounted_str and str(truck_mounted_str).strip().lower() in ['si', 'sí', 'yes']:
            is_truck_mounted = True
            
        # Registramos o actualizamos el Producto
        product, created = Product.objects.update_or_create(
            item_code=str(item_code).strip(),
            defaults={
                'brand': brand,
                'description': str(description).strip(),
                'origin': origin_val,
                'family': brand_name,
                'subfamily': str(subfamily).strip() if subfamily else "",
                'is_truck_mounted': is_truck_mounted
            }
        )
        if created:
            products_created += 1
            
    print(f"Importación completa: {brands_created} marcas creadas, {products_created} productos cargados.")
    return brands_created, products_created
