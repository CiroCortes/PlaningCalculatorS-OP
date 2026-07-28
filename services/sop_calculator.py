import openpyxl
from core.models import PurchaseRequest

class SOPCalculatorService:
    @staticmethod
    def calculate_sop_metrics(file_path):
        """
        Calcula y extrae todas las métricas de S&OP de la pestaña 'Resumen' del Excel,
        consolidándolas con las solicitudes de compra de la Base de Datos, tanto en formato
        global como SKU por SKU (código a código), mapeadas a las columnas correctas de la fila 15.
        """
        wb = openpyxl.load_workbook(file_path, data_only=True)
        if 'Resumen' not in wb.sheetnames:
            raise ValueError("No se encontró la pestaña 'Resumen' en el Excel.")
            
        ws = wb['Resumen']
        
        # 1. --- CÁLCULOS GLOBALES (CONSOLIDADO TOTAL) ---
        delivered_12m_val = ws.cell(row=3, column=3).value or 0.0
        delivered_12m_qty = ws.cell(row=3, column=4).value or 0
        delivered_2026_val = ws.cell(row=4, column=3).value or 0.0
        delivered_2026_qty = ws.cell(row=4, column=4).value or 0
        backlog_val = ws.cell(row=5, column=3).value or 0.0
        backlog_qty = ws.cell(row=5, column=4).value or 0
        cotizado_val = ws.cell(row=6, column=3).value or 0.0
        cotizado_qty = ws.cell(row=6, column=4).value or 0
        budget_val = ws.cell(row=7, column=3).value or 0.0
        budget_qty = ws.cell(row=7, column=4).value or 0
        budget_compliance_qty = ws.cell(row=8, column=3).value or 0.0
        budget_compliance_val = ws.cell(row=9, column=3).value or 0.0
        rotation_months = ws.cell(row=10, column=3).value or 0.0

        stock_val = ws.cell(row=3, column=7).value or 0.0
        stock_qty = ws.cell(row=3, column=8).value or 0
        transit_real_val = ws.cell(row=4, column=7).value or 0.0
        transit_real_qty = ws.cell(row=4, column=8).value or 0
        transit_proj_val = ws.cell(row=5, column=7).value or 0.0
        transit_proj_qty = ws.cell(row=5, column=8).value or 0
        
        # Vaciamos la columna de proceso/solicitud manual del Excel
        excel_process_val = 0.0
        excel_process_qty = 0
        
        backlog_minus_val = ws.cell(row=8, column=7).value or 0.0
        backlog_minus_qty = ws.cell(row=8, column=8).value or 0
        
        # Vaciamos también las solicitudes de compra iniciales del Excel
        excel_purchase_val = 0.0
        excel_purchase_qty = 0

        # Sumamos solicitudes de BD a nivel global (excluyendo rechazadas)
        db_requests = PurchaseRequest.objects.exclude(status='RECHAZADO')
        db_purchase_qty = 0
        db_purchase_val = 0.0
        
        sku_db_map = {}
        for req in db_requests:
            qty = req.quantity
            val = float(req.quantity) * float(req.unit_cost_usd)
            db_purchase_qty += qty
            db_purchase_val += val
            
            sku_str = req.product.item_code.strip()
            if sku_str not in sku_db_map:
                sku_db_map[sku_str] = {'qty': 0, 'val': 0.0}
            sku_db_map[sku_str]['qty'] += qty
            sku_db_map[sku_str]['val'] += val

        # Inyectamos solicitudes de BD en "En Proceso"
        process_qty = excel_process_qty + db_purchase_qty
        process_val = excel_process_val + db_purchase_val

        # Recalculamos Inventario Inicial y Final dinámicamente con las solicitudes reales
        initial_inventory_qty = stock_qty + transit_real_qty + transit_proj_qty + process_qty
        initial_inventory_val = stock_val + transit_real_val + transit_proj_val + process_val

        final_inventory_qty = initial_inventory_qty - backlog_minus_qty
        final_inventory_val = initial_inventory_val - backlog_minus_val

        purchase_qty = excel_purchase_qty + db_purchase_qty
        purchase_val = excel_purchase_val + db_purchase_val
        final_projected_val = final_inventory_val + purchase_val
        final_projected_qty = final_inventory_qty + purchase_qty
        
        monthly_demand_qty = (delivered_12m_qty + backlog_qty) / 12.0
        rotation_with_purchase_months = 0.0
        if monthly_demand_qty > 0:
            rotation_with_purchase_months = final_projected_qty / monthly_demand_qty

        global_metrics = {
            'demand': {
                'delivered_12m_val': round(float(delivered_12m_val), 2),
                'delivered_12m_qty': int(delivered_12m_qty),
                'delivered_2026_val': round(float(delivered_2026_val), 2),
                'delivered_2026_qty': int(delivered_2026_qty),
                'backlog_val': round(float(backlog_val), 2),
                'backlog_qty': int(backlog_qty),
                'cotizado_val': round(float(cotizado_val), 2),
                'cotizado_qty': int(cotizado_qty),
                'budget_val': round(float(budget_val), 2),
                'budget_qty': int(budget_qty),
                'budget_compliance_qty': round(float(budget_compliance_qty) * 100.0, 2),
                'budget_compliance_val': round(float(budget_compliance_val) * 100.0, 2),
                'rotation_months': round(float(rotation_months), 2),
                'rotation_with_purchase_months': round(float(rotation_with_purchase_months), 2)
            },
            'inventory': {
                'stock_val': round(float(stock_val), 2),
                'stock_qty': int(stock_qty),
                'transit_real_val': round(float(transit_real_val), 2),
                'transit_real_qty': int(transit_real_qty),
                'transit_proj_val': round(float(transit_proj_val), 2),
                'transit_proj_qty': int(transit_proj_qty),
                'process_val': round(float(process_val), 2),
                'process_qty': int(process_qty),
                'initial_inventory_val': round(float(initial_inventory_val), 2),
                'initial_inventory_qty': int(initial_inventory_qty),
                'backlog_minus_val': round(float(backlog_minus_val), 2),
                'backlog_minus_qty': int(backlog_minus_qty),
                'final_inventory_val': round(float(final_inventory_val), 2),
                'final_inventory_qty': int(final_inventory_qty),
                'purchase_val': round(float(purchase_val), 2),
                'purchase_qty': int(purchase_qty),
                'final_projected_val': round(float(final_projected_val), 2),
                'final_projected_qty': int(final_projected_qty)
            }
        }

        # 2. --- CÁLCULOS SKU POR SKU (TABLA DETALLADA CON COLUMNAS CORRECTAS DE ROW 15) ---
        sku_details = []
        for r in range(16, 112):
            sku_code = ws.cell(row=r, column=2).value
            if not sku_code:
                continue
            
            sku_str = str(sku_code).strip()
            origin = ws.cell(row=r, column=3).value or ""
            family = ws.cell(row=r, column=4).value or ""
            subfamily = ws.cell(row=r, column=5).value or ""
            desc = ws.cell(row=r, column=6).value or ""
            
            # Stock (Col J / K)
            sku_stock_qty = int(ws.cell(row=r, column=10).value or 0)
            sku_stock_val = float(ws.cell(row=r, column=11).value or 0.0)
            
            # Tránsito Real (Col L / N)
            sku_transit_real_qty = int(ws.cell(row=r, column=12).value or 0)
            sku_transit_real_val = float(ws.cell(row=r, column=14).value or 0.0)
            
            # Tránsito Proyectado (Col O / Q)
            sku_transit_proj_qty = int(ws.cell(row=r, column=15).value or 0)
            sku_transit_proj_val = float(ws.cell(row=r, column=17).value or 0.0)
            
            # Vaciamos la columna de proceso/solicitud manual del Excel
            sku_process_qty = 0
            sku_process_val = 0.0
            
            # Inventario Inicial (Col U / W) -> Calculado dinámicamente con las solicitudes de BD
            db_sku_data = sku_db_map.get(sku_str, {'qty': 0, 'val': 0.0})
            sku_process_qty = int(db_sku_data['qty'])
            sku_process_val = float(db_sku_data['val'])

            sku_init_qty = sku_stock_qty + sku_transit_real_qty + sku_transit_proj_qty + sku_process_qty
            sku_init_val = sku_stock_val + sku_transit_real_val + sku_transit_proj_val + sku_process_val
            
            # Backlog (Col X / Z)
            sku_back_qty = int(ws.cell(row=r, column=24).value or 0)
            sku_back_val = float(ws.cell(row=r, column=26).value or 0.0)
            
            # Inventario Final (Col AA / AC) -> Calculado dinámicamente
            sku_fin_qty = sku_init_qty - sku_back_qty
            sku_fin_val = sku_init_val - sku_back_val
            
            # Solicitud de compras (Col AD / AF) -> Inyectamos las solicitudes de BD
            sku_pur_qty = int(db_sku_data['qty'])
            sku_pur_val = float(db_sku_data['val'])
            
            # Inventario Final Proyectado (Col AG / AI) -> Calculado dinámicamente
            sku_proj_qty = sku_fin_qty + sku_pur_qty
            sku_proj_val = sku_fin_val + sku_pur_val
            
            # Entregados 12M (Col AK / AM)
            sku_del_12m_qty = int(ws.cell(row=r, column=37).value or 0)
            sku_del_12m_val = float(ws.cell(row=r, column=39).value or 0.0)
            
            # Presupuesto (Col 19 / 20)
            sku_budget_qty = ws.cell(row=r, column=19).value or 0
            sku_budget_val = ws.cell(row=r, column=20).value or 0.0
            
            # Rotación Estándar (Row 17 Col W = Col 23) -> Calculada en base a inventario final dinámico
            sku_monthly_demand = (int(sku_del_12m_qty) + int(sku_back_qty)) / 12.0
            sku_rot = 0.0
            if sku_monthly_demand > 0:
                sku_rot = sku_fin_qty / sku_monthly_demand

            # Recalculamos rotación con solicitudes por SKU
            consolidated_rot_pur = 0.0
            if sku_monthly_demand > 0:
                consolidated_rot_pur = sku_proj_qty / sku_monthly_demand

            sku_details.append({
                'code': sku_str,
                'description': desc,
                'family': family,
                'subfamily': subfamily,
                'origin': origin,
                'demand': {
                    'delivered_12m_qty': int(sku_del_12m_qty),
                    'delivered_12m_val': float(sku_del_12m_val),
                    'backlog_qty': int(sku_back_qty),
                    'backlog_val': float(sku_back_val),
                    'budget_qty': int(sku_budget_qty),
                    'budget_val': float(sku_budget_val),
                    'rotation_months': round(float(sku_rot), 2),
                    'rotation_with_purchase_months': round(float(consolidated_rot_pur), 2)
                },
                'inventory': {
                    'stock_qty': int(sku_stock_qty),
                    'stock_val': float(sku_stock_val),
                    'transit_real_qty': int(sku_transit_real_qty),
                    'transit_real_val': float(sku_transit_real_val),
                    'transit_proj_qty': int(sku_transit_proj_qty),
                    'transit_proj_val': float(sku_transit_proj_val),
                    'process_qty': int(sku_process_qty),
                    'process_val': float(sku_process_val),
                    'initial_inventory_qty': int(sku_init_qty),
                    'initial_inventory_val': float(sku_init_val),
                    'backlog_minus_qty': int(sku_back_qty),
                    'backlog_minus_val': float(sku_back_val),
                    'final_inventory_qty': int(sku_fin_qty),
                    'final_inventory_val': float(sku_fin_val),
                    'purchase_qty': int(sku_pur_qty),
                    'purchase_val': float(sku_pur_val),
                    'final_projected_qty': int(sku_proj_qty),
                    'final_projected_val': float(sku_proj_val)
                }
            })

        return {
            'global': global_metrics,
            'skus': sku_details
        }
